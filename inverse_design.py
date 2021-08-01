import numpy as np
import pandas as pd
import openpyxl, time
from skopt import forest_minimize, dummy_minimize, gp_minimize
from skopt.space import Real, Categorical
from skopt.utils import use_named_args
from own_package.pso_ga import pso_ga
from own_package.features_labels_setup import load_data_to_fl, load_testset_to_fl
from own_package.others import print_array_to_excel, create_results_directory, print_df_to_excel, create_excel_file
from own_package.active_learning import load_svm_ensemble, svm_ensemble_prediction, load_model_ensemble, \
    model_ensemble_prediction


def inverse_design(targets, loss_func, bounds, init_guess, model_directory_store, svm_directory, loader_file, write_dir,
                   opt_mode, opt_params):
    '''
    Run inverse design experiment. Give a set of trained model and a target labels, this optimizer determines a list of
    suitable candidate experimental conditions to achieve those target labels.
    :param targets: Targets for the labels
    :param loss_func: Loss function which can be customized according to different logic
    :param bounds: Bounds on the feature search space
    :param init_guess: Initial guess for features. Set as None if nothing.
    :param model_directory_store: list of directories which contain the models used for inverse design
    :param svm_directory: directory that contains the SVM classifier to determine if a composition if feasible or not
    :param loader_file: data loader excel file for the final round used to trained the model. Is used to get the scaler
    for scaling the features
    :param write_dir: directory to write the excel results into
    :param opt_mode: to determine what type of optimizer to use for the inverse design
    :param opt_params: parameters for the optimizer
    1) psoga: Particle swarm, genetic algorithm hybrid optimizer
    2) forest: Forest optimizer from skopt package
    3) dummy: Random search from skopt package
    '''

    model_store = []
    for model_directory in model_directory_store:
        model_store.extend(load_model_ensemble(model_directory))
    svm_store = load_svm_ensemble(svm_directory)
    fl = load_data_to_fl(loader_file, norm_mask=[0, 1, 3, 4, 5], normalise_labels=False)
    data_store = []

    def calculate_score_from_features(features):
        # From features, calculate the score and other results
        x = features[0]
        y = features[1]
        # Ensure that composition sums to 1 by reflecting points across the plane y=1-x from top right to bottom left
        if x + y > 1:
            u = -y + 1
            v = -x + 1
            features[0:2] = np.array([u, v])
        p_class, distance = svm_ensemble_prediction(svm_store, features[0:2])  # SVM Check
        if distance.item() < 0:
            # Distance should be negative value when SVM assigns class 0. Hence a_score will be negative.
            # The more negative the a_score is, the further the composition is from the hyperplane,
            # hence, the less likely the optimizer will select examples with class 0.
            score = -10e5 * distance.item()
            prediction_mean = [-1] * fl.labels_dim
            prediction_std = [-1] * fl.labels_dim
            disagreement = -1
        elif features[0] + features[1] > 1:
            # Distance should be negative value when SVM assigns class 0. Hence a_score will be negative.
            # The more negative the a_score is, the further the composition is from the hyperplane,
            # hence, the less likely the optimizer will select examples with class 0.
            score = -10e5 * (1 - (features[0] + features[1]))
            prediction_mean = [-1] * fl.labels_dim
            prediction_std = [-1] * fl.labels_dim
            disagreement = -1
        else:
            features_c = features[:-1]
            onehot = features[-1].item()
            if onehot == 0:
                features_in = np.concatenate((features_c, np.array([1, 0, 0])))
            elif onehot == 1:
                features_in = np.concatenate((features_c, np.array([0, 1, 0])))
            elif onehot == 2:
                features_in = np.concatenate((features_c, np.array([0, 0, 1])))
            features_input_norm = fl.apply_scaling(features_in)
            prediction_mean, prediction_std = model_ensemble_prediction(model_store, features_input_norm)
            score = loss_func(targets, prediction_mean)
            disagreement = np.mean(prediction_std)
            prediction_mean = prediction_mean.tolist()
            prediction_std = prediction_std.tolist()
        return score, disagreement, prediction_mean, prediction_std

    if opt_mode == 'psoga':
        def fitness(params):
            nonlocal data_store
            features = np.array(params)
            score, disagreement, prediction_mean, prediction_std = calculate_score_from_features(features)
            data = list(features) + [score, disagreement] + prediction_mean + prediction_std
            data_store.append(data)
            return (score,)
        # pso_ga parameters
        pmin = [x[0] for x in bounds]
        pmax = [x[1] for x in bounds]
        smin = [abs(x - y) * 0.001 for x, y in zip(pmin, pmax)]
        smax = [abs(x - y) * 0.5 for x, y in zip(pmin, pmax)]
        # run pso_ga
        pso_ga(func=fitness, pmin=pmin, pmax=pmax,
               smin=smin, smax=smax,
               int_idx=[3], params=opt_params, ga=True, initial_guess=init_guess)
    elif opt_mode == 'forest' or opt_mode == 'dummy':
        # skopt parameters
        space = [Real(low=bounds[0][0], high=bounds[0][1], name='CNT'),
                 Real(low=bounds[1][0], high=bounds[1][1], name='PVA'),
                 Real(low=bounds[2][0], high=bounds[2][1], name='Thickness'),
                 Categorical(categories=[0, 1, 2], name='Dimension')]
        iter_count = 0
        start = time.time()
        end = 0
        @use_named_args(space)
        def fitness(**params):
            nonlocal data_store, iter_count, start, end
            iter_count +=1
            features = np.array([x for x in params.values()])
            score, disagreement, prediction_mean, prediction_std = calculate_score_from_features(features)
            data = list(features) + [score, disagreement] + prediction_mean + prediction_std
            data_store.append(data)
            if iter_count % 10 == 0:
                end = time.time()
                print('Current Iteration {}. Time taken for past 10 evals: {}. '.format(iter_count, end-start))
                start = time.time()
            return score
        # Run skopt optimizer
        if opt_mode == 'gp':
            gp_minimize(func=fitness,
                            dimensions=space,
                            acq_func='EI',  # Expected Improvement.
                            n_calls=opt_params['total_run'],
                            n_random_starts=opt_params['random_run'],
                            verbose=False)
        else:
            dummy_minimize(func=fitness,
                            dimensions=space,
                            n_calls=opt_params['total_run'],
                            verbose=False)

    # Preparing results dataframe
    p_mean_name = ['Pmean_' + str(x) for x in list(map(str, np.arange(1, 4)))]
    p_std_name = ['Pstd_' + str(x) for x in list(map(str, np.arange(1, 4)))]
    columns = fl.features_c_names[:-3].tolist()+['dim','score', 'disagreement']+p_mean_name+p_std_name
    iter_df = pd.DataFrame(data=data_store,
                           columns=columns)
    iter_df = iter_df.sort_values(by=['score'], ascending=True)
    # Print results to excel
    excel_dir = create_excel_file('{}/inverse_design_{}_{}.xlsx'.format(write_dir, opt_mode, targets))
    wb = openpyxl.load_workbook(excel_dir)
    ws = wb[wb.sheetnames[-1]]
    ws.cell(1, 1).value = 'Target'
    print_array_to_excel(array=targets, first_cell=(1, 2), axis=1, ws=ws)
    print_df_to_excel(df=iter_df, ws=ws, start_row=3)
    wb.save(excel_dir)
    wb.close()

