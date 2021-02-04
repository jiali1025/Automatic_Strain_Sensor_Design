import math
import openpyxl
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import pickle
from sklearn.preprocessing import MinMaxScaler
from own_package.features_labels_setup import Features_labels_grid, load_data_to_fl
from own_package.others import print_df_to_excel, create_excel_file


def read_excel_data_to_cutoff(read_excel_file, write_dir):
    '''
    Reads raw data excel file and convert it to labels (GF10,GF100,end) and store in another new excel file
    :param read_excel_file: Directory of raw data excel
    :param write_dir: Directory to save the new excel to
    :return: None
    '''
    cutoff = [10,100]
    # read_excel_file part
    df = pd.read_excel(read_excel_file, sheet_name='raw', header=[0, 1], index_col=0, engine='openpyxl')
    # take only strain columns and make into a new df
    strain = df.xs('Strain (%)', level='Data', axis=1)
    strain = strain.values.T.tolist()  # .T to ensure that each sub list is along the column rather than rows of the df
    # strain store is a list of 1d ndarray, with each inner list being one set of strain data for one experiment
    strain_store = []
    for single_exp in strain:
        strain_single = [x for x in single_exp if not np.isnan(x)]
        strain_store.append(np.array(strain_single))
    # Similar to the above process, but repeat for relative resistance instead
    r = df.xs('R', level='Data', axis=1)
    r = r.values.T.tolist()
    r_store = []
    for single_exp in r:
        r_single = [x for x in single_exp if not np.isnan(x)]
        r_store.append(r_single)

    # Calculation of cutoff for each experiment.
    # Labels are: 1) strain when GF=10, 2) strain when GF=100, 3) strain when sensor breaks
    # Labelled as GF10, GF100, end
    cutoff_store = []
    for strain, r in zip(strain_store, r_store):
        r = np.array(r)
        strain = np.array(strain)
        gf_store = (r[1:] - r[:-1])/(strain[1:] - strain[:-1]) * 100
        if gf_store[-1]<-1:
            cutoff_one = -1
            cutoff_two = -1
        else:
            # Logic for converting the raw data to GF10, GF100, End
            try:
                cut_idx = np.where(gf_store>=cutoff[0])[0][0]
                if strain[cut_idx] > 0:
                    cutoff_one = strain[cut_idx]
                else:
                    cutoff_one = strain[cut_idx+1]
            except IndexError:
                cutoff_one = strain[-1]
            try:
                cut_idx = np.where(gf_store>=cutoff[1])[0][0]
                if strain[cut_idx] > 0:
                    cutoff_two = strain[cut_idx]
                else:
                    cutoff_two = strain[cut_idx+1]
            except IndexError:
                cutoff_two = strain[-1]
        cutoff_store.append([cutoff_one, cutoff_two, strain[-1]])

    # Print to write_excel_file
    excel_name = write_dir + '/results.xlsx'
    wb = openpyxl.Workbook()
    wb.create_sheet('cutoff')
    ws = wb['cutoff']
    # Prepare dataframe and print to excel
    header = np.array(range(np.shape(strain_store)[0])) + 1  # Index to label Exp 1, 2, 3, ...
    columns = ['cut={}'.format(x) for x in cutoff] + ['END']
    header = list(header)
    df_write = pd.DataFrame(cutoff_store, index=header, columns=columns)
    print_df_to_excel(df_write, ws=ws)
    # Save excel
    wb.save(excel_name)
    wb.close()


def read_grid_data(read_excel_file, write_dir):
    '''
    Reads raw gird data excel file and convert it to features and labels in a grid_fl class object and
    store as a pkl file
    :param read_excel_file: Directory of raw data excel
    :param write_dir: Directory to save the new excel to
    :return: None
    '''
    # read_excel_file part
    df = pd.read_excel(read_excel_file, index_col=0, engine='openpyxl')
    df = df.replace('A',1)
    df = df.replace(['B', 'C', 'D'], 0)

    # Converting 2D grid into feature array and labels array
    features = []
    labels = []
    for i in df.index:  # for each row (CNT)
        for j in df.columns.tolist():  # for each column (PVA)
            val = df.loc[i,j]
            if val == 1:
                features.append([i,j])
                labels.append(1)
            elif val == 0:
                features.append([i,j])
                labels.append(0)

    features = np.array(features)/100  # Convert % to fraction
    fl = Features_labels_grid(features=features, labels=labels, idx=None)
    with open(write_dir + '/grid_data', 'wb') as handle:
        pickle.dump(fl, handle, protocol=pickle.HIGHEST_PROTOCOL)
    return fl


def l2_tracker(write_excel_dir, final_excel_loader, last_idx_store):
    '''
    To calculate the average min(L2 distance) over all the data points.
    The avg min L2 is caclulated for each active learnning round, as indicated by the last_idx_store
    :param write_excel_dir: Excel directory to write the data to
    :param final_excel_loader: The excel loader file that contains the feature information
    :param last_idx_store: A list to indicate which experiment number is the last experiment for that batch of active
    learning round. For example, we have 3 active learning rounds with 5, 10, and 3 experiments per round.
    So the last idx store will be [5, 15, 18]

    Saves a new excel file which contains the L2 information
    1) It contains the avg min L2 for each batch of active learning round
    2) The avg min L2 distance for the batch of suggestions for the next active learning round.
    Since the last round has no additional suggestions, the last round has no calculated value for this.
    '''
    write_excel_dir = create_excel_file(write_excel_dir)
    wb = openpyxl.Workbook()
    wb.create_sheet('L2 Results')
    ws = wb[wb.sheetnames[-1]]
    scaler = MinMaxScaler()
    scaler.fit(np.array([[200], [2000]]))
    fl = load_data_to_fl(data_loader_excel_file=final_excel_loader, normalise_labels=False, scaler=scaler,
                         norm_mask=[0, 1, 3, 4, 5])
    final_features = fl.features_c_norm
    suggestions_store = [y2 - y1 for y2, y1 in zip(last_idx_store[1:], last_idx_store[:-1])] + [0]
    batch_l2_store = []
    batch_l2_suggestions_store = []
    for last_idx, suggestions_numel in zip(last_idx_store, suggestions_store):
        features = final_features[:last_idx, :].tolist()

        l2_store = []
        for idx, x in enumerate(features):
            other_features = np.array(features[:idx] + features[idx + 1:])
            l2_distance = np.linalg.norm(x=other_features - np.array(x).reshape((1, -1)), ord=2, axis=1)
            l2_store.append(np.min(l2_distance))
        batch_l2_store.append(np.mean(l2_store))

    df = pd.DataFrame(data=np.concatenate((np.array(last_idx_store).reshape(-1, 1),
                                           np.array(batch_l2_store).reshape(-1, 1),), axis=1),
                      columns=['Expt Batches', 'Mean Min L2'],
                      index=range(1, len(last_idx_store) + 1))
    print_df_to_excel(df=df, ws=ws)
    wb.save(write_excel_dir)


