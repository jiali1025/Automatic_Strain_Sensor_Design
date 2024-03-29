import pandas as pd
import numpy as np
import pickle, time, os
import openpyxl
from openpyxl import load_workbook
from skopt import gp_minimize, dummy_minimize
from skopt.space import Integer, Real
from skopt.utils import use_named_args
# Own Scripts
from own_package.models import create_hparams
from .cross_validation import run_kf
from own_package.others import print_array_to_excel, create_excel_file, print_df_to_excel


def hparam_opt(model_mode, fl, fl_store, other_fl_dict, scoring, total_run, write_dir, random_run=10, plot_dir=None):
    data_store_dir = write_dir + '/data_store'
    run_count = 0
    data_store = []

    if model_mode == 'ann':
        # Prepare bounds for search
        bounds = [[10, 100, ],
                 [50, 1000]]
        #bounds = [[10, 30, ],
        #          [10, 50]]
        nodes = Integer(low=bounds[0][0], high=bounds[0][1], name='nodes')
        epochs = Integer(low=bounds[1][0], high=bounds[1][1], name='epochs')
        dimensions = [nodes, epochs]
        default_parameters = [20, 50]
        data_store_count = 1
        data_store_name = 0

        # Fitness function to evaluate the score for each trial of hyperparameters
        @use_named_args(dimensions=dimensions)
        def fitness(nodes, epochs):
            nonlocal run_count, data_store, fl, fl_store, data_store, data_store_count, data_store_name
            start_time = time.time()
            run_count += 1
            # run_kf for current trial of hyperparameters and return the score
            hparams = create_hparams(nodes=nodes, epochs=epochs, loss=scoring, learning_rate=0.001,
                                     reg_l1=0.0005, reg_l2=0, verbose=0)
            if plot_dir:
                plot_name = '{}/{}_{}_run_{}'.format(plot_dir, model_mode, scoring, run_count)
            else:
                plot_name = None
            val_score, results_dict = run_kf(model_mode=model_mode, fl=fl, fl_store=fl_store, hparams=hparams,
                                             scoring=scoring, other_fl_dict=other_fl_dict, write_excel_dir=None,
                                             save_model_name=f'{write_dir}/models/{scoring}_{model_mode}_run{run_count}',
                                             plot_name=plot_name)
            results_dict['info']['opt'] = {'nodes': nodes, 'epochs': epochs}
            results_dict['info']['model_name'] = f'{write_dir}_run{run_count}'
            # Save results
            if (data_store_count - 1) % 5 == 0:
                data_store = []
                data_store_name += 5
            data_store.append(results_dict)
            with open('{}/data_store_{}.pkl'.format(data_store_dir, data_store_name), "wb") as file:
                pickle.dump(data_store, file)
            data_store_count += 1
            end_time = time.time()
            print(
                f'**************************************************************************************************\n'
                f'Run Number {run_count} \n'
                f'nodes: {nodes}, epochs: {epochs}\n'
                f'Time Taken: {end_time - start_time}\n'
                f'*********************************************************************************************')
            return val_score
    elif model_mode == 'dtr' or model_mode == 'dtrc':
        # Prepare bounds for search
        if model_mode == 'dtrc':
            chain=True
        else:
            chain=False
        bounds = [[1, 10, ],
                 [1, 1000]]
        #bounds = [[1, 5, ],
        #          [1, 10]]
        depth = Integer(low=bounds[0][0], high=bounds[0][1], name='depth')
        num_est = Integer(low=bounds[1][0], high=bounds[1][1], name='num_est')
        dimensions = [depth, num_est]
        default_parameters = [[5,5]] #[[462,30],
                              #[438,4],
                              #[391,488]]
        data_store_count = 1
        data_store_name = 0

        @use_named_args(dimensions=dimensions)
        def fitness(depth, num_est):
            nonlocal run_count, data_store, fl, fl_store, data_store_count, data_store_name
            start_time = time.time()
            run_count += 1
            # run_kf for single trial of hyperparameter
            hparams = create_hparams(max_depth=depth, num_est=num_est, chain=chain)
            val_score, results_dict = run_kf(model_mode=model_mode, fl=fl, fl_store=fl_store, hparams=hparams,
                                             scoring=scoring, other_fl_dict=other_fl_dict, write_excel_dir=None,
                                             save_model_name=f'{write_dir}/models/{scoring}_{model_mode}_run{run_count}',
                                             plot_name=None)
            results_dict['info']['opt'] = {'depth': depth, 'num_est': num_est}
            results_dict['info']['model_name'] = f'{write_dir}_run{run_count}'
            # Save results in batches
            if (data_store_count - 1) % 5 == 0:
                data_store = []
                data_store_name += 5
            data_store.append(results_dict)
            # Save data_store batch every trial in case hparam_opt accidentally terminates early (e.g. server shut down)
            with open('{}/data_store_{}.pkl'.format(data_store_dir, data_store_name), "wb") as file:
                pickle.dump(data_store, file)
            data_store_count += 1
            end_time = time.time()
            print(f'*************************************************************************************************\n'
                  f'Run Number {run_count} \n'
                  f'Depth {depth}, No. Estimators {num_est}\n'
                  f'Time Taken: {end_time - start_time}\n'
                  f'*********************************************************************************************')
            return val_score

    elif model_mode == 'svr':
        # Prepare bounds for search
        bounds = [[-4, 2],
                 [-1, 3]]
        #bounds = [[1, 5, ],
        #          [1, 10]]
        gamma = Real(low=bounds[0][0], high=bounds[0][1], name='gamma')
        C = Real(low=bounds[1][0], high=bounds[1][1], name='C')
        dimensions = [gamma, C]
        default_parameters = [-2, 0]
        data_store_count = 1
        data_store_name = 0

        @use_named_args(dimensions=dimensions)
        def fitness(gamma, C):
            nonlocal run_count, data_store, fl, fl_store, data_store_count, data_store_name
            start_time = time.time()
            run_count += 1
            # run_kf for single trial of hyperparameter
            hparams = create_hparams(gamma=float(10.0)**gamma, C=float(10.0)**C)
            val_score, results_dict = run_kf(model_mode=model_mode, fl=fl, fl_store=fl_store, hparams=hparams,
                                             scoring=scoring, other_fl_dict=other_fl_dict, write_excel_dir=None,
                                             save_model_name=f'{write_dir}/models/{scoring}_{model_mode}_run{run_count}',
                                             plot_name=None)
            results_dict['info']['opt'] = {'gamma': 10.0**gamma, 'C': 10.0**C}
            results_dict['info']['model_name'] = f'{write_dir}_run{run_count}'
            # Save results in batches
            if (data_store_count - 1) % 5 == 0:
                data_store = []
                data_store_name += 5
            data_store.append(results_dict)
            # Save data_store batch every trial in case hparam_opt accidentally terminates early (e.g. server shut down)
            with open('{}/data_store_{}.pkl'.format(data_store_dir, data_store_name), "wb") as file:
                pickle.dump(data_store, file)
            data_store_count += 1
            end_time = time.time()
            print(f'*************************************************************************************************\n'
                  f'Run Number {run_count} \n'
                  f'Gamma {10.0**gamma}, C {10.0**C}\n'
                  f'Time Taken: {end_time - start_time}\n'
                  f'*********************************************************************************************')
            return val_score



    search_result = gp_minimize(func=fitness,
                                dimensions=dimensions,
                                acq_func='EI',  # Expected Improvement.
                                n_calls=total_run,
                                n_random_starts=random_run,
                                x0=default_parameters)

    # Print hyperparameter optimization summary results into excel
    wb = load_workbook(write_dir + '/hparam_results.xlsx')
    hparam_store = np.array(search_result.x_iters)
    results = np.array(search_result.func_vals)
    index = np.arange(total_run) + 1
    toprint = np.concatenate((index.reshape(-1, 1), hparam_store, results.reshape(-1, 1)), axis=1)
    if model_mode == 'ann':
        header = np.array(['index', 'nodes', 'epochs', 'mse'])
    elif model_mode == 'dtr':
        header = np.array(['index', 'max_depth', 'num_est', 'mse'])
    elif model_mode == 'svr':
        header = np.array(['index', 'gamma', 'C', 'mse'])
    toprint = np.concatenate((header.reshape(1, -1), toprint), axis=0)
    sheetname = wb.sheetnames[-1]
    ws = wb[sheetname]
    print_array_to_excel(toprint, (1, 1), ws, axis=2)
    wb.save(write_dir + '/hparam_results.xlsx')
    wb.close()


def read_hparam_opt_data_store(write_dir):
    # Load all the saved data_store.pkl into data_store list
    data_store = []
    for filename in os.listdir(f'{write_dir}/data_store'):
        if filename.endswith(".pkl"):
            with open('{}/{}'.format(f'{write_dir}/data_store', filename), 'rb') as handle:
                data_store.extend(pickle.load(handle))

    wb_store = {k: openpyxl.Workbook() for k in data_store[0] if (k not in ['info'])}
    summary_df = []
    for run, data in enumerate(data_store):
        run_summary = data['info']['opt']
        for k, wb in wb_store.items():
            wb.create_sheet(f'run {run + 1}')
            ws = wb[f'run {run + 1}']
            print_df_to_excel(df=data[k]['df'], ws=ws)
            print_df_to_excel(df=data['info']['hparams'], ws=ws, start_col=len(data[k]['df'].columns) + 3)
            run_summary = {**run_summary, **{'model_name': data['info']['model_name'],f'{k}_mse': data[k]['mse'], f'{k}_mre': data[k]['mre']}}
        summary_df.append(pd.Series(run_summary))
    # Print summary df
    summary_df = pd.concat(summary_df, axis=1).T
    wb = openpyxl.Workbook()
    ws = wb[wb.sheetnames[-1]]
    print_df_to_excel(df=summary_df, ws=ws)
    wb.save(f'{write_dir}/summary.xlsx')
    wb.close()
    # Print all the other df
    for k, wb in wb_store.items():
        wb.save(f'{write_dir}/{k}.xlsx')
        wb.close()



def combine_data_store(dir_store):

    top_df = []
    name_store = []
    for dir in dir_store:
        data_store = []
        for filename in os.listdir(f'{dir}/data_store'):
            if filename.endswith(".pkl"):
                with open('{}/data_store/{}'.format(dir, filename), 'rb') as handle:
                    data_store.extend(pickle.load(handle))
        ett_store = np.array([data['val']['mse'] for data in data_store])

        top_idx = np.argsort(ett_store)[:3]
        name_store.extend([data_store[idx]['info']['model_name'] for idx in top_idx])
        top_df.extend([data_store[idx]['ett30_pointsend']['df'] for idx in top_idx])
    print(name_store)
    wb = openpyxl.Workbook()
    for idx, df in enumerate(top_df):
        wb.create_sheet(f's{idx}')
        ws = wb[f's{idx}']
        print_df_to_excel(df,ws)
    wb.create_sheet('Avg')
    ws = wb['Avg']
    df = pd.concat(top_df).reset_index().groupby('index').mean()
    print_df_to_excel(df,ws)
    p_y = df.iloc[:,-19:].values
    y = df.iloc[:,-39:-20].values
    print( np.mean((p_y - y) ** 2), np.mean(np.abs(p_y - y).T / y[:, -1]))
    ws = wb['Sheet']
    ws.cell(1,1,'mse')
    ws.cell(1, 2, 'mre')
    ws.cell(2,1, np.mean((p_y - y) ** 2))
    ws.cell(2, 2, np.mean(np.abs(p_y - y).T / y[:, -1]))
    wb.save(f'./results/top3.xlsx')
    wb.close()


