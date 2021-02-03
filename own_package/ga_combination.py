import pandas as pd
import numpy as np
import openpyxl, pickle
import matplotlib.pyplot as plt
import os, itertools, pickle
from collections import defaultdict
from deap import algorithms, base, creator, tools
from sklearn.metrics import mean_squared_error
from own_package.cross_validation import mean_relative_error
from own_package.others import print_array_to_excel, print_df_to_excel, create_results_directory


def prepare_grand_data_store(dir_store):
    data_store = []
    for dir in dir_store:
        for filename in os.listdir(f'{dir}/data_store'):
            if filename.endswith(".pkl"):
                with open('{}/data_store/{}'.format(dir, filename), 'rb') as handle:
                    data_store.extend(pickle.load(handle))
    return data_store


def ga_opt(load_dir_store, hparams):
    # Load all the saved data_store.pkl into data_store list
    data_store = prepare_grand_data_store(load_dir_store)

    yt = data_store[0]['train']['df'].iloc[:, -6:-3].values
    p_yt_store = np.array([data['train']['df'].iloc[:, -3:].values for data in data_store])
    yv = data_store[0]['val']['df'].iloc[:, -6:-3].values
    p_yv_store = np.array([data['val']['df'].iloc[:, -3:].values for data in data_store])

    def eval(individual):
        # Individual is a list of 0 or 1, where if the j entry is 1, the j model is included and vice versa
        selected_mask = [idx for idx, value in enumerate(individual) if value == 1]
        # Calculate mean haitao error for the selected models
        re_t = mean_haitao_error(yt, np.mean(p_yt_store[selected_mask, :, :], axis=0))
        re_v = mean_haitao_error(yv, np.mean(p_yv_store[selected_mask, :, :], axis=0))
        re = (re_t + 2 * re_v) / 3
        return (re,)

    creator.create("FitnessMax", base.Fitness, weights=(-1,))
    creator.create("Individual", list, fitness=creator.FitnessMax)
    toolbox = base.Toolbox()
    toolbox.register("attr_bool", np.random.choice, np.arange(0, 2), p=hparams['init'])
    toolbox.register("individual", tools.initRepeat, creator.Individual, toolbox.attr_bool, n=len(data_store))
    toolbox.register("population", tools.initRepeat, list, toolbox.individual)
    toolbox.register("evaluate", eval)
    toolbox.register("mate", tools.cxTwoPoint)
    toolbox.register("mutate", tools.mutFlipBit, indpb=0.2)
    toolbox.register("select", tools.selTournament, tournsize=3)
    # Logging
    stats = tools.Statistics(key=lambda ind: ind.fitness.values)
    stats.register("avg", np.mean, axis=0)
    stats.register("std", np.std, axis=0)
    stats.register("min", np.min, axis=0)
    stats.register("max", np.max, axis=0)
    pop = toolbox.population(n=hparams['n_pop'])
    hof = tools.HallOfFame(1)
    # Run the GA algorithm
    pop, logbook = algorithms.eaSimple(toolbox=toolbox, population=pop,
                                       cxpb=0.5, mutpb=0.2,
                                       ngen=hparams['n_gen'], halloffame=hof, stats=stats,
                                       verbose=True)

    # Create the ga results dir based on the load dir name
    results_dir = create_results_directory(f'./results/ga/ga_opt',
                                           folders=['plots'], excels=['ga_results'])
    # Plotting
    gen = logbook.select("gen")
    fit_min = [x.item() for x in logbook.select("min")]
    fit_avg = [x.item() for x in logbook.select("avg")]
    fit_max = [x.item() for x in logbook.select("max")]
    fig, ax1 = plt.subplots()
    line1 = ax1.plot(gen, fit_min, label="Min MRE")
    line2 = ax1.plot(gen, fit_avg, label="Avg MRE")
    line3 = ax1.plot(gen, fit_max, label="Max MRE")
    plt.legend()
    ax1.set_xlabel("Generation")
    ax1.set_ylabel("Haitao Error")
    plt.savefig('{}/plots/GA_opt_MRE_all.png'.format(results_dir), bbox_inches="tight")
    fig, ax1 = plt.subplots()
    line1 = ax1.plot(gen, fit_min, label="Min MRE")
    plt.legend()
    ax1.set_xlabel("Generation")
    ax1.set_ylabel("Total Generation Cost")
    plt.savefig('{}/plots/GA_opt_min_only.png'.format(results_dir), bbox_inches="tight")

    # Store final results
    av = hof[-1]  # av stands for allocation vector
    results_dict = defaultdict(list)
    data_names = [k for k in data_store[0].keys() if k not in ['info']]
    for data, indicator in zip(data_store, av):
        if indicator == 1:  # Means include the model
            for k in data_names:
                results_dict[k].append(data[k]['df'].iloc[:, -3:].values)

    # Create excel workbook to print GA results to
    wb = openpyxl.Workbook()
    # Print allocation vector to excel
    wb.create_sheet('av')
    ws = wb['av']
    model_names = [data['info']['model_name'] for data in data_store]
    print_df_to_excel(df=pd.DataFrame([av, model_names], index=['av', 'model_names']).T, ws=ws)
    summary_df = {}
    for k, v in results_dict.items():  # Print the prediction for each dataset to excel
        y = data_store[0][k]['df'].iloc[:, -6:-3].values
        v = np.array(v)
        p_y = np.mean(v, axis=0)
        mse = mean_squared_error(y, p_y)
        mre = mean_relative_error(y, p_y)
        var = np.mean(np.var(v, axis=0))
        summary_df[k] = {'mse': mse, 'mre': mre, 'var': var}
        df = pd.DataFrame(np.hstack((y, p_y)), columns=[f'y{i + 1}' for i in range(3)] + [f'P_y{i + 1}' for i in
                                                                                          range(3)])
        wb.create_sheet(k)
        ws = wb[k]
        print_df_to_excel(df=df, ws=ws)
        print_df_to_excel(df=pd.DataFrame.from_dict({'mse': [mse],
                                                     'mre': [mre]}), ws=ws, start_col=10)
    # Print summary of losses for different dataset in the summary worksheet
    summary_df = pd.DataFrame.from_dict(summary_df)
    def move_column_inplace(df, col, pos):
        col = df.pop(col)
        df.insert(pos, col.name, col)
    move_column_inplace(summary_df, 'train', 0)
    move_column_inplace(summary_df, 'val', 1)
    ws = wb['Sheet']
    print_df_to_excel(df=summary_df, ws=ws, start_row=5)
    print_df_to_excel(df=pd.DataFrame(hparams), ws=ws)
    # Save and close excel workbook
    wb.save(f'{results_dir}/ga_results.xlsx')
    wb.close()
