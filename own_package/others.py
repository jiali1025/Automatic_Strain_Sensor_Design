import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl
import sys
import os

def create_results_directory(results_directory, folders=None, excels=None):
    if os.path.exists(results_directory):
        expand = 1
        while True:
            expand += 1
            new_results_directory = results_directory + ' - ' + str(expand)
            if os.path.exists(new_results_directory):
                continue
            else:
                results_directory = new_results_directory
                break

    os.mkdir(results_directory)

    if folders:
        for item in folders:
            os.mkdir(results_directory + '/' + item)

    if excels:
        for item in excels:
            if item[-5:] != '.xlsx':
                item = item + '.xlsx'
            excel_name = results_directory + '/' + item
            wb = openpyxl.Workbook()
            wb.save(excel_name)
            wb.close()

    print('Creating new results directory: {}'.format(results_directory))
    return results_directory


def print_array_to_excel(array, first_cell, ws, axis=2):
    '''
    Print an np array to excel using openpyxl
    :param array: np array
    :param first_cell: first cell to start dumping values in
    :param ws: worksheet reference. From openpyxl, ws=wb[sheetname]
    :param axis: to determine if the array is a col vector (0), row vector (1), or 2d matrix (2)
    '''
    if isinstance(array, (list,)):
        array = np.array(array)
    shape = array.shape
    if axis == 0:
        # Treat array as col vector and print along the rows
        array=array.flatten()  # Flatten in case the input array is a nx1 ndarry which acts weird
        for i in range(shape[0]):
            j = 0
            ws.cell(i + first_cell[0], j + first_cell[1]).value = array[i]
    elif axis == 1:
        # Treat array as row vector and print along the columns
        array=array.flatten()  # Flatten in case the input array is a 1xn ndarry which acts weird
        for j in range(shape[0]):
            i = 0
            ws.cell(i + first_cell[0], j + first_cell[1]).value = array[j]
    elif axis == 2:
        # If axis==2, means it is a 2d array
        for i in range(shape[0]):
            for j in range(shape[1]):
                ws.cell(i + first_cell[0], j + first_cell[1]).value = array[i, j]

def print_df_to_excel(df, ws, start_row=1, start_col=1, index=True, header=True):
    rows = list(dataframe_to_rows(df, index=index, header=header))
    rows.pop(1)
    for r_idx, row in enumerate(rows, start_row):
        skip_count = 0
        for c_idx, value in enumerate(row, start_col):
            if isinstance(value, str):
                if 'Unnamed' not in value:
                    ws.cell(row=r_idx - skip_count, column=c_idx, value=value)
            else:
                ws.cell(row=r_idx - skip_count, column=c_idx, value=value)
        else:
            skip_count += 1


def create_excel_file(excel_name):
    while os.path.isfile(excel_name):
        expand = 1
        while True:
            expand += 1
            new_file_name = excel_name.split('.xlsx')[0] + ' - ' + str(expand) + '.xlsx'
            if os.path.isfile(new_file_name):
                continue
            else:
                excel_name = new_file_name
                break
    print('Writing into' + excel_name + '\n')
    wb = openpyxl.Workbook()
    wb.save(excel_name)
    return excel_name


def round_sigfigs(num, sig_figs):
    """Round to specified number of sigfigs.

    round_sigfigs(0, sig_figs=4)
    0
    int(round_sigfigs(12345, sig_figs=2))
    12000
    int(round_sigfigs(-12345, sig_figs=2))
    -12000
    int(round_sigfigs(1, sig_figs=2))
    1
    '{0:.3}'.format(round_sigfigs(3.1415, sig_figs=2))
    '3.1'
    '{0:.3}'.format(round_sigfigs(-3.1415, sig_figs=2))
    '-3.1'
    '{0:.5}'.format(round_sigfigs(0.00098765, sig_figs=2))
    '0.00099'
    '{0:.6}'.format(round_sigfigs(0.00098765, sig_figs=3))
    '0.000988'
    """
    if num != 0:
        return round(num, -int(math.floor(math.log10(abs(num))) - (sig_figs - 1)))
    else:
        return 0  # Can't take the log of 0

if __name__ == '__main__':
    pass