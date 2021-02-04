from tensorflow.python.keras import backend as K
from tensorflow.python.keras.models import load_model
import numpy as np
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os, time, gc, pickle, itertools, math
from typing import List
import time

from skopt import gp_minimize, gbrt_minimize, forest_minimize, dummy_minimize
from skopt.space import Real, Categorical
from skopt.utils import use_named_args
from skopt.plots import plot_convergence
from sklearn.metrics import pairwise_distances
from own_package.features_labels_setup import load_data_to_fl
from own_package.others import print_array_to_excel, create_results_directory, print_df_to_excel, create_excel_file
from own_package.pso_ga import pso_ga


def load_svm_ensemble(model_directory) -> List:
    """
    Load list of trained svm models from a pickle saved file that can be used for prediction later
    :param model_directory: model directory where the h5 models are saved in. NOTE: All the models in the directory will
     be loaded. Hence, make sure all the models in there are part of the ensemble and no unwanted models are in the
     directory
    :return: [List: svm models]
    """
    # Loading model names into a list
    model_name_store = []
    directory = model_directory
    for idx, file in enumerate(os.listdir(directory)):
        filename = os.fsdecode(file)
        model_name_store.append(directory + '/' + filename)
    print('Loading the following models from {}. Total models = {}'.format(directory, len(model_name_store)))
    # Loading model class object into a list
    model_store = []
    for name in model_name_store:
        with open(name, "rb") as input_file:
            model_store.append(pickle.load(input_file))
        print('Model {} has been loaded'.format(name))

    return model_store


def svm_ensemble_prediction(model_store, composition, probability=False):
    """
    Run prediction given one set of feactures_c_norm input, using all the models in model store.
    :param model_store: List of keras models returned by the def load_model_ensemble
    :param Composition: ndarray of shape (-1, 2). The columns represents the CNT and PVA composition.
    :return: List of metrics.
    """
    predictions_store = []
    distance_store = []
    proba_store = []
    if len(composition.shape) == 1:
        composition = composition.reshape(1, 2)
    for model in model_store:
        predictions_store.append(model.model.predict(composition))
        distance_store.append(model.model.decision_function(composition))
        if probability:
            proba_store.append(model.model.predict_proba(composition)[:, 1])

    predictions = np.round(np.average(np.array(predictions_store), axis=0), decimals=0)
    distance = np.mean(np.array(distance_store), axis=0)

    if probability:
        probability = np.mean(np.array(proba_store), axis=0)
        return predictions, distance, probability
    else:
        return predictions, distance


def relative_error(y_true, y_pred):
    diff = K.abs((y_true - y_pred) / K.reshape(K.clip(K.abs(y_true[:, -1]),
                                                      K.epsilon(),
                                                      None), (-1, 1)))
    return 100. * K.mean(diff, axis=-1)


def load_model_ensemble(model_directory) -> List:
    """
    Load list of trained keras models from a .h5 saved file that can be used for prediction later
    :param model_directory: model directory where the h5 models are saved in. NOTE: All the models in the directory will
     be loaded. Hence, make sure all the models in there are part of the ensemble and no unwanted models are in the
     directory
    :return: [List: keras models]
    """
    # Loading model names into a list
    model_name_store = []
    directory = model_directory
    for idx, file in enumerate(os.listdir(directory)):
        filename = os.fsdecode(file)
        model_name_store.append(directory + '/' + filename)
    print('Loading the following models from {}. Total models = {}'.format(directory, len(model_name_store)))
    # Loading model class object into a list
    model_store = []
    for name in model_name_store:
        if name.endswith(".pkl"):  # DTR models
            model_store.append(pickle.load(open(name, 'rb')))
        elif name.endswith('.h5'):  # ANN models
            try:
                model_store.append(load_model(name))
            except ValueError:
                model_store.append(load_model(name, compile=False))
        else:
            print('{} found that does not end with .pkl or .h5'.format(name))
        print('Model {} has been loaded'.format(name))

    return model_store


def model_ensemble_prediction(model_store, features_c_norm):
    """
    Run prediction given one set of feactures_c_norm input, using all the models in model store.
    :param model_store: List of models returned by the def load_model_ensemble
    :param features_c_norm: ndarray of shape (1, -1). The columns represents the different features.
    :return: List of metrics.
    """
    predictions_store = []
    for model in model_store:
        p_y = model.predict(features_c_norm)
        for row, p_label in enumerate(p_y.tolist()):
            if p_label[1] > p_label[2]:
                p_y[row, 1] = p_y[row, 2]
            if p_label[0] > p_y[row, 1]:
                p_y[row, 0] = p_y[row, 1]
        predictions_store.append(p_y)
    predictions_store = np.array(predictions_store).squeeze()
    predictions_mean = np.mean(predictions_store, axis=0)
    predictions_std = np.std(predictions_store, axis=0)
    return predictions_mean, predictions_std


def acquisition_opt(bounds, svm_directory, loader_file, normalise_labels, write_dir, opt_mode, opt_params,
                    batch_runs=1, ignore_distance=False, norm_mask=None):
    '''
    To perform batch-wise active learning for each round.
    :param bounds: Features search space
    :param svm_directory: Directory that contains the SVM models
    :param loader_file: fl excel data loader
    :param normalise_labels: for fl
    :param write_dir: Directory to write excel to and also where the model directory is in
    :param opt_mode: Choose the type of optimizer
    :param opt_params: Parameters for optimizer
    :param batch_runs: Number of batches of experiments to run
    :param ignore_distance: When calculating acquisition score, whether to consider L2 distance or not
    :param norm_mask: for fl
    '''
    # Load models from latest round
    model_store = load_model_ensemble(f'{write_dir}/models')
    svm_store = load_svm_ensemble(svm_directory)
    # Load latest round of fl class
    fl = load_data_to_fl(loader_file, norm_mask=norm_mask, normalise_labels=normalise_labels)
    excel_file = create_excel_file(f'{write_dir}/{opt_mode}_acq.xlsx')
    wb = openpyxl.Workbook()

    def calculate_score_from_features(features):
        x = features[0]
        y = features[1]
        if x + y > 1:
            u = -y + 1
            v = -x + 1
            features[0:2] = np.array([u, v])

        # SVM Check
        p_class, distance = svm_ensemble_prediction(svm_store, features[0:2])
        if distance.item() < 0:
            # Distance should be negative value when SVM assigns class 0. Hence a_score will be negative.
            # The more negative the a_score is, the further the composition is from the hyperplane,
            # hence, the less likely the optimizer will select examples with class 0.
            a_score = 10e5 * distance.item()
            prediction_mean = [-1] * fl.labels_dim
            prediction_std = [-1] * fl.labels_dim
            l2_distance = -1
            disagreement = -1
        elif features[0] + features[1] > 1:
            # Sum of composition cannot be greater than 1
            a_score = 10e5 * (1 - (features[0] + features[1]))
            prediction_mean = [-1] * fl.labels_dim
            prediction_std = [-1] * fl.labels_dim
            l2_distance = -1
            disagreement = -1
        else:
            features_c = features[:-1]
            onehot = features[-1].item()
            if onehot == 0:
                features = np.concatenate((features_c, np.array([1, 0, 0])))
            elif onehot == 1:
                features = np.concatenate((features_c, np.array([0, 1, 0])))
            elif onehot == 2:
                features = np.concatenate((features_c, np.array([0, 0, 1])))

            features_input_norm = fl.apply_scaling(features)
            prediction_mean, prediction_std = model_ensemble_prediction(model_store, features_input_norm)
            prediction_mean = prediction_mean.tolist()
            prediction_std = prediction_std.tolist()
            # Greedy Sampling
            # Get L2 distance of sampled example to all existing example in fl class object
            # Note: L2 distance is calculated using the normalised features so that all feature have the same weight
            l2_distance = np.linalg.norm(x=fl.features_c_norm - features_input_norm.reshape((1, -1)), ord=2, axis=1)
            l2_distance = np.min(l2_distance)  # Take the minimum L2 dist.
            # Overall Acquisition Score. Higher score if l2 distance is larger and uncertainty (std) is larger.
            disagreement = np.sum(prediction_std)
            if ignore_distance:
                a_score = disagreement
            else:
                a_score = l2_distance * disagreement
        return a_score, l2_distance, disagreement, prediction_mean, prediction_std

    for batch in range(batch_runs):
        instance_start = time.time()
        iter_count = 0
        data_store = []
        if opt_mode in ['gp', 'dummy', 'forest']:
            # skopt parameters setup
            space = [Real(low=bounds[0][0], high=bounds[0][1], name='CNT'),
                     Real(low=bounds[1][0], high=bounds[1][1], name='PVA'),
                     Real(low=bounds[2][0], high=bounds[2][1], name='Thickness'),
                     Categorical(categories=[0, 1, 2], name='Dimension')]

            @use_named_args(space)
            def fitness(**params):
                nonlocal iter_count, data_store
                iter_count += 1
                features = np.array([x for x in params.values()])
                a_score, l2_distance, disagreement, prediction_mean, prediction_std = calculate_score_from_features(
                    features)
                # Storing intermediate results into list to print into excel later
                data = list(features) + [a_score, disagreement, l2_distance] + prediction_mean + prediction_std
                data_store.append(data)
                if iter_count % 50 == 0:
                    print(f'Current Iteration: {iter_count} out of {opt_params["total_run"]} for batch {batch + 1}.')
                return -a_score  # -ve to maximise the a_score

            if opt_mode == 'gp':
                search_result = gp_minimize(func=fitness,
                                            dimensions=space,
                                            acq_func='EI',  # Expected Improvement.
                                            n_calls=opt_params['total_run'],
                                            n_random_starts=opt_params['random_run'],
                                            verbose=False)
            elif opt_mode == 'dummy':
                search_result = dummy_minimize(func=fitness,
                                               dimensions=space,
                                               n_calls=opt_params['total_run'],
                                               verbose=False)
            elif opt_mode == 'forest':
                search_result = forest_minimize(func=fitness,
                                                dimensions=space,
                                                acq_func='EI',  # Expected Improvement.
                                                n_calls=opt_params['total_run'],
                                                n_random_starts=opt_params['random_run'],
                                                verbose=False)
            best_x = search_result.x
        elif opt_mode == 'psoga':
            # psoga parameters setup
            pmin = [x[0] for x in bounds]
            pmax = [x[1] for x in bounds]
            smin = [abs(x - y) * 0.001 for x, y in zip(pmin, pmax)]
            smax = [abs(x - y) * 0.5 for x, y in zip(pmin, pmax)]

            def fitness(params):
                nonlocal data_store
                features = np.array(params)
                a_score, l2_distance, disagreement, prediction_mean, prediction_std = calculate_score_from_features(
                    features)
                data = list(features) + [a_score, disagreement, l2_distance] + prediction_mean + prediction_std
                data_store.append(data)
                return (-a_score,)

            _, _, best_x = pso_ga(func=fitness, pmin=pmin, pmax=pmax,
                                  smin=smin, smax=smax,
                                  int_idx=[3], params=opt_params, ga=True, initial_guess=None)
        else:
            raise TypeError(f'Invalid opt_mode {opt_mode}')

        # Prepare results dataframe
        p_mean_name = ['Pmean_' + str(x) for x in list(map(str, np.arange(1, 4)))]
        p_std_name = ['Pstd_' + str(x) for x in list(map(str, np.arange(1, 4)))]
        columns = fl.features_c_names[:-3].tolist() + ['dim', 'A_score', 'disagreement',
                                                       'L2'] + p_mean_name + p_std_name
        iter_df = pd.DataFrame(data=data_store, columns=columns)
        iter_df = iter_df.sort_values(by=['A_score'], ascending=False)
        # Creating new worksheet.
        wb.create_sheet(title='Batch_{}'.format(batch + 1))
        ws = wb['Batch_{}'.format(batch + 1)]
        print_df_to_excel(df=iter_df, ws=ws)
        '''
        If more than one batch, prepare fl for next batch. The only difference is that the previous best trial point
        with the highest a_score will be added to fl.features_c_norm such that the L2 greedy distance will
        account for the fact that the previous batch would had contained the best example already.
        '''
        features = np.array(best_x)
        features_c = features[:-1]
        onehot = features[-1].item()
        if onehot == 0:
            features = np.concatenate((features_c, np.array([1, 0, 0])))
        elif onehot == 1:
            features = np.concatenate((features_c, np.array([0, 1, 0])))
        elif onehot == 2:
            features = np.concatenate((features_c, np.array([0, 0, 1])))
        fl.features_c_norm = np.concatenate((fl.features_c_norm, fl.apply_scaling(features)), axis=0)

        instance_end = time.time()
        print('Batch {} completed. Time taken: {}'.format(batch + 1, instance_end - instance_start))
        wb.save(excel_file)


def l2_points_opt(numel, write_dir, svm_directory, seed_number_of_expt, total_expt, l2_opt=True):
    write_dir = create_results_directory(results_directory=write_dir, excels=['l2_acq'])
    svm_store = load_svm_ensemble(svm_directory)
    base = [x / (numel * 2 - 1) for x in list(range(numel * 2))]

    # Create set of possible compositions
    compositions = [[x, y] if x + y <= 1 else [-x + 1, -y + 1] for x, y in
                    list(itertools.product(base[::2], base[1::2]))]
    distance_store = []
    # Check feasibility for those compositions
    for model in svm_store:
        distance_store.append(model.model.decision_function(compositions))
    distance = np.mean(np.array(distance_store), axis=0)
    valid_compositions = [x for x, dist in zip(compositions, distance) if dist >= 0]
    print('Number of compositions = {}. % valid = {}%'.format(len(valid_compositions),
                                                              len(valid_compositions) / len(compositions) * 100))
    # Permute feasible compositions with different thickness possibilities scaled from 0 to 1
    number_valid_compositions = round(math.sqrt(len(valid_compositions)))
    compositions_thickness = list(itertools.product(valid_compositions,
                                                    [x / (number_valid_compositions - 1)
                                                     for x in list(range(number_valid_compositions))]))
    print('Number of permutations = {}'.format(len(compositions_thickness * 3)))
    # Permute the above with 0D, 1D, and 2D
    all_permutations = np.array([x[0] + [x[1]] + y
                                 for x in compositions_thickness for y in [[1, 0, 0], [0, 1, 0], [0, 0, 1]]])

    if l2_opt:
        expt_idx = np.random.randint(0, len(all_permutations), seed_number_of_expt)
        expt_store = all_permutations[expt_idx, :]

        for i in range(total_expt - seed_number_of_expt):
            start = time.time()
            d = pairwise_distances(expt_store, all_permutations, metric='euclidean')
            next_expt = np.argmax(np.min(d, axis=0))
            expt_store = np.concatenate((expt_store, all_permutations[next_expt, None, :]), axis=0)
            end = time.time()
            print(
                '{} out of {} completed. Time taken = {}.'.format(i + 1, total_expt - seed_number_of_expt, end - start))
    else:
        expt_idx = np.random.randint(0, len(all_permutations), total_expt)
        expt_store = all_permutations[expt_idx, :]

    expt_store[:, 2] = expt_store[:, 2] * 1800 + 200

    write_excel = '{}/l2_acq.xlsx'.format(write_dir)
    wb = openpyxl.load_workbook(write_excel)
    wb.create_sheet('l2_acq')
    ws = wb[wb.sheetnames[-1]]
    ws.cell(1, 1).value = 'Valid Combinations'
    ws.cell(1, 2).value = len(all_permutations)
    ws.cell(1, 3).value = 'Seed Expt'
    ws.cell(1, 4).value = seed_number_of_expt
    df = pd.DataFrame(data=expt_store, columns=['CNT', 'PVA', 'Thickness', '0D', '1D', '2D'],
                      index=list(range(1, total_expt + 1)))
    print_df_to_excel(df=df, ws=ws, start_row=2)

    wb.save(write_excel)
    pass
