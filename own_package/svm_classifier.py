from sklearn.svm import SVC
import pandas as pd
import numpy as np
import gc, pickle, time
import openpyxl
from skopt import gp_minimize, dummy_minimize
from skopt.space import Real, Integer
from skopt.utils import use_named_args
from skopt.plots import plot_convergence
from sklearn.metrics import matthews_corrcoef, mean_squared_error
# Own Scripts
# from own_package.svr import  run_svr
from own_package.others import print_array_to_excel, create_excel_file, print_df_to_excel


class SVMmodel:
    def __init__(self, fl, gamma=1):
        # Model setup
        self.features_dim = fl.features_dim
        self.labels_dim = fl.labels_dim  # Assuming that each task has only 1 dimensional output
        self.model = SVC(kernel='rbf', gamma=gamma, probability=True)  # Set up SVM radial basis model

    def train_model(self, fl):
        training_features = fl.features
        training_labels = fl.labels
        self.model.fit(training_features, training_labels)
        return self

    def predict(self, eval_fl):
        features = eval_fl.features
        y_pred = self.model.predict(features)
        return y_pred


def run_classification(grid_fl_dir, write_dir, gamma):
    # Load grid fl
    with open(grid_fl_dir, 'rb') as handle:
        fl = pickle.load(handle)
    # Create 10 fold for cross validation
    fl_store = fl.create_kf(k_folds=10, shuffle=True)
    # Run k model instance to perform skf
    # Results dataframe has the columns: ['idx', 'fold', 'CNT', 'PVA', 'Label', 'Prediction']
    # For each fold, append the fold information to the following lists:
    val_idx = []
    folds = []
    val_features = []
    val_labels = []
    predicted_labels_store = []
    # fl_store is a 10 item list where each item is a tuple containing the train and val fl
    for fold, fl_tuple in enumerate(fl_store):
        instance_start = time.time()
        (ss_fl, i_ss_fl) = fl_tuple  # ss_fl is training fl, i_ss_fl is validation fl
        # Train model
        model = SVMmodel(fl=ss_fl, gamma=gamma)
        model.train_model(fl=ss_fl)
        # Evaluation
        predicted_labels = model.predict(i_ss_fl)
        # Saving model
        save_model_name = write_dir + '/models/svm_' + str(fold + 1) + '.pkl'
        print('Saving instance {} model in {}'.format(fold + 1, save_model_name))
        with open(save_model_name, 'wb') as handle:
            pickle.dump(model, handle, protocol=pickle.HIGHEST_PROTOCOL)
        # Preparing data to put into new_df that consists of all the validation dataset and its predicted labels
        val_idx.extend(i_ss_fl.idx)
        folds.extend([fold] * i_ss_fl.count)  # Make a col that contains the fold number for each example
        if len(val_features):
            val_features = np.concatenate((val_features, i_ss_fl.features), axis=0)
        else:
            val_features = i_ss_fl.features
        val_labels.extend(i_ss_fl.labels)
        predicted_labels_store.extend(predicted_labels)
        # Printing one instance summary.
        instance_end = time.time()
        print(
            '\nFor k-fold run {} out of {}. Each fold has {} examples. Time taken for '
            'instance = {}\n'
            '####################################################################################################'
                .format(fold + 1, 10, i_ss_fl.count, instance_end - instance_start))

    # Calculating metrics based on complete validation prediction
    mcc = matthews_corrcoef(y_true=val_labels, y_pred=predicted_labels_store)

    # Creating dataframe to print into excel later.
    results_df = np.concatenate((np.array(folds)[:, None],  # Convert 1d list to col. vector
                             val_features,
                             np.array(val_labels)[:, None],
                             np.array(predicted_labels_store)[:, None]) , axis=1)
    headers = ['folds'] + \
              ['CNT', 'PVA'] + \
              ['Labels'] + \
              ['Prediction']
    # val_idx is the original position of the example in the data_loader
    results_df = pd.DataFrame(data=results_df, columns=headers, index=val_idx)
    # Create excel file and print results to excel
    excel_file = create_excel_file(f'{write_dir}/classifier_results.xlsx')
    print('Writing into' + excel_file)
    wb = openpyxl.Workbook()
    # Create results sheet
    wb.create_sheet('results')
    ws = wb['results']
    # Print results df
    print_df_to_excel(df=results_df, ws=ws)
    # Writing hyperparameter information at the side
    start_col = len(results_df.columns) + 3
    headers = ['mcc', 'gamma']
    values = [mcc, gamma]
    print_array_to_excel(np.array(headers), (1, start_col + 1), ws, axis=1)
    print_array_to_excel(np.array(values), (2, start_col + 1), ws, axis=1)
    wb.save(excel_file)
    wb.close()


def svm_hparam_opt(grid_fl_dir, total_run, write_excel_dir):
    with open(grid_fl_dir, 'rb') as fp:
        fl = pickle.load(fp)

    run_count = 0
    gamma = Real(low=0.1, high=300, name='gamma')
    dimensions = [gamma]
    default_parameters = [130]

    fl_store = fl.create_kf(k_folds=10, shuffle=True)

    @use_named_args(dimensions=dimensions)
    def fitness(gamma):
        nonlocal run_count, fl_store
        run_count += 1
        # Run k model instance to perform skf
        predicted_labels_store = []
        val_labels = []
        for fold, fl_tuple in enumerate(fl_store):
            (ss_fl, i_ss_fl) = fl_tuple  # ss_fl is training fl, i_ss_fl is validation fl

            # Training
            model = SVMmodel(fl=ss_fl, gamma=gamma)
            model.train_model(fl=ss_fl)

            # Evaluation
            predicted_labels = model.predict(i_ss_fl).flatten().tolist()
            predicted_labels_store.extend(predicted_labels)
            val_labels.extend(i_ss_fl.labels.flatten().tolist())

        # Calculating metrics based on complete validation prediction
        mcc = matthews_corrcoef(y_true=val_labels, y_pred=predicted_labels_store)
        if run_count % 10 == 0:  # Print every 10 iteration
            print(f'Run Number {run_count}')
        return -mcc

    search_result = gp_minimize(func=fitness,
                                dimensions=dimensions,
                                acq_func='EI',  # Expected Improvement.
                                n_calls=total_run,
                                x0=default_parameters)
    print('Best Loss = {}'.format(search_result.fun))
    print('Best Gamma = {}'.format(search_result.x[0]))
    x = [x[0] for x in search_result.x_iters]
    results = pd.DataFrame([x] + [(-search_result.func_vals).tolist()]).T
    results.columns = ['Gamma', 'mcc']
    results = results.sort_values(by='mcc', ascending=False)

    write_excel_dir = create_excel_file(write_excel_dir)
    wb = openpyxl.load_workbook(write_excel_dir)
    ws = wb[wb.sheetnames[-1]]
    print_df_to_excel(df=results, ws=ws)
    wb.save(write_excel_dir)
    wb.close()
