import tensorflow as tf
from tensorflow.python.keras import backend as K
import gc
import numpy as np
import pandas as pd
from pandas import Series
import openpyxl
from openpyxl import load_workbook
from sklearn.metrics import mean_squared_error
import time, pickle

# Own Scripts
from own_package.models import Kmodel, DTRmodel
from .others import print_array_to_excel, print_df_to_excel


def mean_haitao_error(y_true, y_pred):
    try:
        return np.mean(np.abs(y_true - y_pred).T / y_true[:, -1])
    except TypeError:
        y_true = np.array(y_true)
        y_pred = np.array(y_pred)
        return np.mean(np.abs(y_true - y_pred).T / y_true[:, -1])


def eval_model_on_fl(model, fl, label_scaler=None):
    p_y = model.predict(fl)
    if label_scaler:  # Means that the prediction returned by the model is normalized. Change to original scale.
        p_y = label_scaler.inverse_transform(p_y)
    # Since it might be possible that the model predicts a strain for GF10 > GF100, then set GF10=GF100,
    # Same for GF100 vs end.
    for row, p_label in enumerate(p_y.tolist()):  # Loop through each experiment (row of data)
        if p_label[1] > p_label[2]:
            p_y[row, 1] = p_y[row, 2]
        if p_label[0] > p_y[row, 1]:
            p_y[row, 0] = p_y[row, 1]
    return p_y


def run_kf(model_mode, fl, fl_store,
           hparams, scoring,
           other_fl_dict=None,
           write_excel_dir=None,
           save_model_name=None,
           plot_name=None):
    '''
    Stratified k fold cross validation for training and evaluating model 2 only. Model 1 data is trained before hand.
    :param model_mode: Choose between using SNN or cDNN (non_smiles) and SNN_smiles or cDNN_smiles
    :param cv_mode: Cross validation mode. Either 'skf' or 'loocv'.
    :param hparams: hparams dict containing hyperparameters information
    :param loader_file: data_loader excel file location
    :param skf_file: skf_file name to save excel file as
    :param skf_sheet: name of sheet to save inside the skf_file excel. If None, will default to SNN or cDNN as name
    :param k_folds: Number of k folds. Used only for skf cv_mode
    :param k_shuffle: Whether to shuffle the given examples to split into k folds if using skf
    :return:
    '''
    valid_scoring = ['mse', 'he']
    if scoring not in valid_scoring:
        raise TypeError(f'Scoring function {scoring} is not a valid option. Choose one of {valid_scoring}')
    k_folds = len(fl_store)
    # Run k model instance to perform skf
    results_dict = {'val': {'df': {'idx': [], 'fold': [], 'features': [], 'labels': [], 'predictions': []},
                            'mse': -1, 'he': -1}}
    if other_fl_dict:
        other_fl_dict = {**other_fl_dict, **{'train': fl}}
        results_dict = {**results_dict,
                        **{k: {'df': {'features': v.features_c, 'labels': v.labels, 'predictions': []},
                               'mse': -1, 'he': -1} for k, v in other_fl_dict.items()}}

    for fold, fl_tuple in enumerate(fl_store):  # Train, eval, save model, store results for each fold
        instance_start = time.time()
        (ss_fl, i_ss_fl) = fl_tuple  # ss_fl is training fl, i_ss_fl is validation fl
        # Set up model
        if model_mode == 'ann':
            sess = tf.compat.v1.Session()
            # sess = tf.Session()
            K.set_session(sess)
            model = Kmodel(fl=ss_fl, hparams=hparams)
        elif model_mode == 'dtr':
            model = DTRmodel(fl=ss_fl, max_depth=hparams['max_depth'], num_est=hparams['num_est'])
        else:
            raise KeyError(f'model_mode {model_mode} is not a valid selection.')
        # Train model and save model training loss vs epoch plot if plot_name is given, else no plot will be saved
        if plot_name:
            model.train_model(ss_fl, i_ss_fl, plot_name='{}_fold_{}.png'.format(plot_name, fold))
        else:
            model.train_model(ss_fl, i_ss_fl)
        # Get prediction
        predicted_labels = eval_model_on_fl(model, fl=i_ss_fl, label_scaler=fl.labels_scaler)
        # Get prediction on the other_fl_store if needed
        if other_fl_dict:
            pass
        # Saving model
        if save_model_name:
            # Set save_model_name
            save_model_name1 = f'{save_model_name}_{fold + 1}'
            # Save model
            print(f'Saving instance {fold + 1} model in {save_model_name1}.h5')
            if model_mode == 'ann':
                model.model.save(save_model_name1 + '.h5')
            elif model_mode == 'dtr':
                pickle.dump(model.model, open(save_model_name1 + '.pkl', 'wb'))
        # Store results in list to convert to results_df later
        results_dict['val']['df']['idx'].extend(i_ss_fl.idx)
        results_dict['val']['df']['fold'].extend([fold] * i_ss_fl.count)
        results_dict['val']['df']['features'].append(i_ss_fl.features_c)
        results_dict['val']['df']['labels'].append(i_ss_fl.labels)
        results_dict['val']['df']['predictions'].append(predicted_labels)
        if other_fl_dict:
            for k, other_fl in other_fl_dict.items():
                other_fl_predictions = eval_model_on_fl(model=model, fl=other_fl, label_scaler=fl.labels_scaler)
                results_dict[k]['df']['predictions'].append(other_fl_predictions)
        # Printing one instance summary.
        instance_end = time.time()
        print(
            f'\nFor k-fold run {fold + 1} out of {k_folds}. Each fold has {i_ss_fl.count} examples.'
            f' Model is {model_mode}. Time taken for '
            f'instance = {instance_end - instance_start}\n'
            '####################################################################################################')
        # Need to put the next few lines if not memory will run out
        del model
        if model_mode == 'ann':
            K.clear_session()
            sess.close()
        gc.collect()

    # Calculate average scores across the folds
    mse_avg = mean_squared_error(np.vstack(results_dict['val']['df']['labels']),
                                 np.vstack(results_dict['val']['df']['predictions']))
    he_avg = mean_haitao_error(np.vstack(results_dict['val']['df']['labels']),
                               np.vstack(results_dict['val']['df']['predictions']))
    results_dict['val']['mse'] = mse_avg
    results_dict['val']['he'] = he_avg

    def get_results_df(results, feature_names, label_names, idx_fold=False):
        # Convert results_dict df stored results into a dataframe
        if idx_fold:
            data = np.hstack((np.array([results['fold']]).T, np.vstack(results['features']),
                              np.vstack(results['labels']), np.vstack(results['predictions'])))
            predicted_label_names = [f'P_{x}' for x in label_names]
            headers = ['folds'] + list(map(str, feature_names)) + list(fl.labels_names) + predicted_label_names
            return pd.DataFrame(data, columns=headers, index=results['idx']).sort_index(axis=0, ascending=True)
        else:
            data = np.hstack((np.vstack(results['features']),
                              np.vstack(results['labels']), np.vstack(results['predictions'])))
            predicted_label_names = [f'P_{x}' for x in label_names]
            headers = list(map(str, feature_names)) + list(fl.labels_names) + predicted_label_names
            return pd.DataFrame(data, columns=headers)

    results_dict['val']['df'] = get_results_df(results=results_dict['val']['df'], feature_names=fl.features_c_names,
                                               label_names=fl.labels_names, idx_fold=True)
    if other_fl_dict:
        for k, other_fl in other_fl_dict.items():
            results_dict[k]['df']['predictions'] = np.mean(np.array(results_dict[k]['df']['predictions']), axis=0)
            results_dict[k]['mse'] = mean_squared_error(results_dict[k]['df']['labels'],
                                                        results_dict[k]['df']['predictions'])
            results_dict[k]['he'] = mean_haitao_error(results_dict[k]['df']['labels'],
                                                      results_dict[k]['df']['predictions'])
            results_dict[k]['df'] = get_results_df(results=results_dict[k]['df'], feature_names=fl.features_c_names,
                                                   label_names=fl.labels_names, idx_fold=False)

    results_dict['info'] = {'hparams': pd.DataFrame(hparams, index=[0])}

    if write_excel_dir:
        # Printing results_df to excel file
        print('Writing into' + write_excel_dir)
        wb = load_workbook(write_excel_dir)
        # Creating new worksheet. Even if results worksheet already exists, a new resultss1 ws will be created and so on
        wb.create_sheet(model_mode)
        ws = wb[wb.sheetnames[-1]]  # Taking the ws name from the back ensures that the new ws is selected
        # Writing results and hparam df
        print_df_to_excel(df=results_dict['val']['df'], ws=ws, start_row=1, start_col=1)
        print_df_to_excel(df=results_dict['info']['hparams'], ws=ws, start_row=1,
                          start_col=len(results_dict['val']['df'].columns) + 3)
        # Writing scores df
        headers = ['mse', 'he']
        values = [[mse_avg, he_avg]]
        scores_df = pd.DataFrame(values, columns=headers, index=['Avg'])
        print_df_to_excel(df=scores_df, ws=ws, start_row=5, start_col=len(results_dict['val']['df'].columns) + 3)
        wb.save(write_excel_dir)
        wb.close()

    if scoring == 'mse':
        return mse_avg, results_dict
    elif scoring == 'he':
        return he_avg, results_dict
